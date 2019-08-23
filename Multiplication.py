import openpyxl

N=sys.argv[1]
wb=openpyxl.Workbook()
sheet=wb.active

for i in range(2,N+2):
    sheet.cell(row=1,column=i).value=i-1

for r in range(2,N+2):
    sheet.cell(row=r,column=1).value=r-1
    for c in range(2,N+2):
        num1=sheet['A'+str(r)].value
        num2=sheet.cell(row=1,column=c).value
        sheet.cell(row=r,column=c).value=num1*num2

wb.save('MuliplicationtTable.xlsx')
print('done')
